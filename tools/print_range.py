from pathlib import Path
from openpyxl import load_workbook

wb_path = Path(__file__).parent.parent.joinpath('Dashboard Projetos - Agência de Inovação UPE.xlsx')
wb = load_workbook(wb_path, data_only=True)
ws = wb['Dashboard']
for r in range(1, 13):
    row_vals = []
    for col in ['AF','AG','AH','AI','AJ','AK','AL']:
        row_vals.append((col+str(r), ws[f"{col}{r}"].value))
    print(row_vals)
wb.close()
