from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

wb_path = Path(__file__).parent.parent.joinpath("Dashboard Projetos - Agência de Inovação UPE.xlsx")
print("Workbook:", wb_path)

# sheet names
from openpyxl import load_workbook
wb = load_workbook(wb_path, data_only=True)
print("Sheets:", wb.sheetnames)
wb.close()

# Dados dos Projetos
print('\n--- Dados dos Projetos ---')
df = pd.read_excel(wb_path, sheet_name='Dados dos Projetos')
df.columns = [str(c).strip() if c is not None else f'Col_{i}' for i, c in enumerate(df.columns)]
print('rows,cols:', df.shape)
print('unique instituições (nao null):', int(df['Instituição'].nunique(dropna=True)) if 'Instituição' in df.columns else 'no column')
print('projetos notna:', int(df['Projeto'].notna().sum()) if 'Projeto' in df.columns else 'no column')

if 'Receita' in df.columns:
    df['Receita'] = pd.to_numeric(df['Receita'], errors='coerce').fillna(0)
    print('receita sum (Dados dos Projetos):', df['Receita'].sum())
else:
    print('Receita column not found')

# Dashboard ranges AF/AG
print('\n--- Dashboard AF/AG (AF1:AF10 / AG1:AG10) ---')
wb = load_workbook(wb_path, data_only=True, read_only=True)
sheet = 'Dashboard'
if sheet in wb.sheetnames:
    ws = wb[sheet]
    anos = []
    receitas = []
    for r in range(1, 11):
        a = ws[f'AF{r}'].value
        b = ws[f'AG{r}'].value
        anos.append(a)
        receitas.append(b if b is not None else 0)
    print('AF anos:', anos)
    print('AG receitas:', receitas)
    s = sum([x for x in receitas if isinstance(x, (int,float))])
    print('sum(AG):', s)
else:
    print('Dashboard sheet not found')
wb.close()

# Also check any total cell commonly used (search for 'Receita Total' text)
print('\n--- Buscar celula com texto RECEITA ---')
wb = load_workbook(wb_path, data_only=True)
found = []
for s in wb.sheetnames:
    ws = wb[s]
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=50, values_only=True):
        for cell in row:
            if isinstance(cell, str) and 'receita' in cell.lower():
                found.append((s, cell))
wb.close()
print('cells with receita keyword (up to 200x50):', found[:20])

print('\nDone')
