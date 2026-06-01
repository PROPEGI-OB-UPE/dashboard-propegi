from pathlib import Path
from openpyxl import load_workbook

target = 9000376.82
wb_path = Path(__file__).parent.parent.joinpath('Dashboard Projetos - Agência de Inovação UPE.xlsx')
wb = load_workbook(wb_path, data_only=True)
coords = []
for s in wb.sheetnames:
    ws = wb[s]
    for row in ws.iter_rows(min_row=1, max_row=300, min_col=1, max_col=100):
        for cell in row:
            val = cell.value
            if isinstance(val, (int, float)):
                if abs(val - target) < 0.001:
                    coords.append((s, cell.coordinate, val))
wb.close()
print('found coords:', coords)
