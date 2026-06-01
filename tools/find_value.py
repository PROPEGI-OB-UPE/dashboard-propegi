from pathlib import Path
from openpyxl import load_workbook

target = 9000376.82
wb_path = Path(__file__).parent.parent.joinpath('Dashboard Projetos - Agência de Inovação UPE.xlsx')
wb = load_workbook(wb_path, data_only=True)
matches = []
for s in wb.sheetnames:
    ws = wb[s]
    for r in ws.iter_rows(min_row=1, max_row=300, min_col=1, max_col=100, values_only=True):
        for cell in r:
            if isinstance(cell, (int, float)):
                if abs(cell - target) < 1.0:  # within 1 real
                    matches.append((s, cell))
wb.close()
print('matches near target:', matches)
if not matches:
    # try relative scale (maybe thousands or formatted differently)
    alt = []
    for s in wb.sheetnames:
        ws = load_workbook(wb_path, data_only=True)[s]
        # handled above, skip for speed
    print('no exact matches found')
